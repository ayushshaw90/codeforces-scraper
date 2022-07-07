from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from codeforces import api
from api_info import key,secret
from openpyxl import Workbook, load_workbook
import sys

#get user modes
print("Welcome to Scraper")
print("Enter 1 for comprehensive scan. It takes longer time and scans all the user data.\n")
print("Enter 2 for quick scan. It checks for users which are newly added friends. It takes less time but gives somewhat outdated results.")
choice = int(input())

if choice !=1 and choice !=2:
    #debug
    print("exited bug")
    sys.exit()


users = (api.call("user.friends", key=key, secret = secret))

# selenium
chrome_options = Options()
chrome_options.add_argument("--headless")
chrome_options.add_argument("--start-maximized")
driver = webdriver.Chrome(options=chrome_options)

data=[]
if choice == 2:
    #debug
    print("entered choice 2")
    map={}
    wb=False
    ws1=0
    try:
        print("entered try block")
        wb=load_workbook('questions_solved.xlsx')
        ws1=wb["stats"]
        run=True
        count=1
        while run:
            count+=1
            if ws1["B"+str(count)].value==None:
                run=False
            else:
                map[ws1["A"+str(count)].value] = ws1["B"+str(count)].value
            
        print(map)
        for i in range(len(users)):
            print("entered inner loop!")
            if not users[i] in map:
                try:
                    driver.get("https://codeforces.com/profile/"+users[i])
                    content = driver.find_element(By.CSS_SELECTOR, 'div._UserActivityFrame_counterValue')
                    dat = content.text
                    l = len(dat)
                    num = int(dat[0:(l-8)])
                    map[users[i]]=str(num)
                    ws1["A"+str(count)]=users[i]
                    ws1["B"+str(count)]=num
                    count+=1
                except NameError:
                    print(NameError)
                    print(users[i]+" caused problem skipping...")
        wb.save('questions_solved.xlsx')
    except:
        if not wb:
            choice = 1
            print("question_solved.xlsx file not found. Switching to full search")
        print("Error working with xlsx document")
    if choice==2:
        sys.exit()

if choice == 1:
    for i in range(len(users)):
        driver.get("https://codeforces.com/profile/"+users[i])
        try:
            content = driver.find_element(By.CSS_SELECTOR, 'div._UserActivityFrame_counterValue')
            dat = content.text
            l = len(dat)
            num = int(dat[0:(l-8)])
            data.append({'solved':num, 'profile':users[i]})
        except:
            print("[-] User "+users[i]+" caused error, skipping...\n")

    driver.close()
    # Sorting the users by number of questions solved
    def sortFunc(e):
        return e['solved']


    data.sort(reverse=True,key = sortFunc)


#Getting output to xlsx file
wb=0
ws1=0
try:
    wb= load_workbook('questions_solved.xlsx')
    ws1=wb.active
except:
    wb=Workbook()
    ws1=wb.active
    ws1.title = 'stats'

ws1['A1'] = "Username"
ws1['B1'] = "Questions solved"
for i in range(len(data)):
    ws1["A"+str(i+2)] = data[i]['profile']
    ws1["B"+str(i+2)] = data[i]['solved']
wb.save('questions_solved.xlsx')